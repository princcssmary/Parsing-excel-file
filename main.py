import openpyxl
from openpyxl import Workbook
from dataclasses import dataclass

SOURCE_FILE = 'Прайс В Мелодия Света - 22_12_2021.xlsx'
TARGET_FILE = 'Файл_куда_нужно_переносить_раздел_Шаблон_Поставщика.xlsx'
SOURCE_WORKSHEET = 'TDSheet'
TARGET_WORKSHEET = 'Шаблон для поставщика'
TYPE_OF_ITEM = 'Люстра ЭкономГудс'
BRAND = 'Нет бренда'
NAME = 'Люстра ЭкономГудс'
PARAM_HEIGHT = 'Коробка Высота'
PARAM_LENGTH = 'Коробка Длина'
PARAM_WIDTH = 'Коробка Ширина'
WEIGHT = 2000
COEFFICIENT_PRICE = 2.5
COEFFICIENT_PRICE_WITHOUT_DISCOUNT = 3
COEFFICIENT_PRICE_OZON_PREMIUM = 2

wb = Workbook()

source_file = openpyxl.load_workbook(SOURCE_FILE)
source_worksheet = source_file[SOURCE_WORKSHEET]

target_file = openpyxl.load_workbook(TARGET_FILE)
target_worksheet = target_file[TARGET_WORKSHEET]


@dataclass
class TargetRow:
    article_number: str
    price: float
    price_without_discount: float
    price_ozon_premium: float
    width: float
    height: float
    length: float
    article_number_of_photo: str
    annotation: list
    type_of_item: str = TYPE_OF_ITEM
    brand: str = BRAND
    name: str = NAME
    weight: int = WEIGHT


@dataclass
class SourceRow:
    article_number: str
    price: int
    annotation: str


def price(data, coefficient):
    if data:
        data = float(data)
        price = data * coefficient
        return price


def parameter(information, param):
    if information:
        information = information.split('\n')
        parameter = []
        for elem in information:
            if param in elem:
                elem = float(elem.split()[2].replace(',', '.')) * 10
                parameter.append(elem)
                for el in parameter:
                    return el


def annotation(information):
    if information:
        information = information.split('\n')
        for elem in information:
            if 'Количество штук в заводской коробке' in elem:
                information.remove(elem)
    return information


sourse_row_list_with_none = []
for i in range(10, source_worksheet.max_row):
    item = SourceRow(article_number=source_worksheet.cell(row=i, column=3).value,
                     price=source_worksheet.cell(row=i, column=6).value,
                     annotation=source_worksheet.cell(row=i, column=11).value)
    sourse_row_list_with_none.append(item)

sourse_row_list = []
for elem in sourse_row_list_with_none:
    if elem.article_number != None:
        sourse_row_list.append(elem)

target_row_list = []
for index in range(0, len(sourse_row_list)):
    t = TargetRow(article_number=sourse_row_list[index].article_number,
                  price=price(sourse_row_list[index].price, COEFFICIENT_PRICE),
                  price_without_discount=price(sourse_row_list[index].price, COEFFICIENT_PRICE_WITHOUT_DISCOUNT),
                  price_ozon_premium=price(sourse_row_list[index].price, COEFFICIENT_PRICE_OZON_PREMIUM),
                  width=parameter(sourse_row_list[index].annotation, PARAM_WIDTH),
                  height=parameter(sourse_row_list[index].annotation, PARAM_HEIGHT),
                  length=parameter(sourse_row_list[index].annotation, PARAM_LENGTH),
                  article_number_of_photo=sourse_row_list[index].article_number,
                  annotation=annotation(sourse_row_list[index].annotation))
    target_row_list.append(t)

for i in range(0, len(target_row_list)):
    target_worksheet.cell(row=i + 4, column=2).value = target_row_list[i].article_number
    target_worksheet.cell(row=i + 4, column=4).value = target_row_list[i].price
    target_worksheet.cell(row=i + 4, column=5).value = target_row_list[i].price_without_discount
    target_worksheet.cell(row=i + 4, column=6).value = target_row_list[i].price_ozon_premium
    target_worksheet.cell(row=i + 4, column=15).value = target_row_list[i].article_number
    target_worksheet.cell(row=i + 4, column=3).value = target_row_list[i].name
    target_worksheet.cell(row=i + 4, column=11).value = target_row_list[i].weight
    target_worksheet.cell(row=i + 4, column=19).value = target_row_list[i].brand
    target_worksheet.cell(row=i + 4, column=21).value = target_row_list[i].type_of_item
    target_worksheet.cell(row=i + 4, column=12).value = target_row_list[i].width
    target_worksheet.cell(row=i + 4, column=13).value = target_row_list[i].height
    target_worksheet.cell(row=i + 4, column=14).value = target_row_list[i].length
    el = None
    if target_row_list[i].annotation:
        el = "\n".join(target_row_list[i].annotation)
    target_worksheet.cell(row=i + 4, column=22).value = el

num = []
for i in range(1, len(target_row_list)+1):
    num.append(i)
for elem in num:
    for i in range(0, len(num)):
        target_worksheet.cell(row=i + 4, column=1).value = num[i]

target_file.save(TARGET_FILE)
