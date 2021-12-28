import openpyxl
from openpyxl import Workbook

NAME = 'Люстра ЭкономГудс'
WEIGHT = 2000
BRAND = 'Нет бренда'
TYPE_OF_ITEM = 'Люстра ЭкономГудс'
ARTICLE_NUMBER_OF_PHOTO = []
ARTICLE_NUMBER = []
PRICE = []
PRICE_WITHOUT_DISCOUNT = []
PRICE_OZON_PREMIUM = []
WIDTH = []
HEIGHT = []
LENGTH = []
ANNOTATION = []
INITIAL_FILE = 'Файл_из_котрого_нужно_переносить.xlsx'
NEW_FILE = 'Файл_куда_нужно_переносить_раздел_Шаблон_Поставщика.xlsx'

wb = Workbook()

excel_file = openpyxl.load_workbook(INITIAL_FILE)
basic_sheet = excel_file['TDSheet']

wb_new = openpyxl.load_workbook(NEW_FILE)

worksheet = wb_new['Шаблон для поставщика']


def list_with_article_numbers():
    """
    Forms list with article numbers of items.
    :rtype: List
    :return: List with article numbers.
    """
    article_number_with_none = [basic_sheet.cell(row=i, column=3).value for i in range(9, 150)]
    article_number = []
    for element in article_number_with_none:
        if element:
            article_number.append(element)
    return article_number


def list_price_of_item(coefficient):
    """
    Forms list with prices of items in accordance with coefficient.
    :param coefficient: Coefficient forming price.
    :rtype: List
    :return: List with prices of items.
    """
    price = [basic_sheet.cell(row=i, column=6).value for i in range(9, 150)]
    price_of_item = []
    for i in price:
        if i != None:
            i = int(i) * coefficient
            price_of_item.append(i)
    return price_of_item


def information_about_item():
    """
    Forms list with all information about items.
    :rtype: List
    :return: List with information about items.
    """
    information = [basic_sheet.cell(row=i, column=11).value for i in range(9, 150)]
    info_of_item = []
    for elem in information:
        if elem:
            elem = elem.split('\n')
            info_of_item.append(elem)
    return info_of_item


def parameters_of_item_in_sm(param):
    """
    Forms list with parameters of item in centimeters.
    :param param: Parameter (width, height, length)
    :rtype: List
    :return: List with information in sm.
    """
    info_of_item = information_about_item()
    list_of_params = []
    for elem in info_of_item:
        if elem:
            for i in elem:
                if param in i:
                    list_of_params.append(i)
    return list_of_params


def parameters_of_item_in_mm(list_of_params_in_sm):
    """
    Forms list with parameters of item in millimeters.
    :param list_of_params_in_sm: List with parameters in sm.
    :rtype: List
    :return: List with parameters (width, height, length) in mm.
    """
    list_of_params_in_mm = []
    for elem in list_of_params_in_sm:
        if elem:
            parameter_in_mm = float(elem.split()[2].replace(',', '.')) * 10
            list_of_params_in_mm.append(parameter_in_mm)
    return list_of_params_in_mm


def annotation():
    """
    Forms list with all characteristics of items without quantity in a box.
    :rtype: List
    :return: List of annotations.
    """
    info_of_item = information_about_item()
    for elem in info_of_item:
        if elem:
            for i in elem:
                if 'Количество штук в заводской коробке' in i:
                    elem.remove(i)
    return info_of_item


if __name__ == '__main__':
    ARTICLE_NUMBER = list_with_article_numbers()
    PRICE = list_price_of_item(2.5)
    PRICE_WITHOUT_DISCOUNT = list_price_of_item(3)
    PRICE_OZON_PREMIUM = list_price_of_item(2)
    ARTICLE_NUMBER_OF_PHOTO = ARTICLE_NUMBER
    ANNOTATION = annotation()
    width_sm = parameters_of_item_in_sm("Коробка Ширина")
    height_sm = parameters_of_item_in_sm("Коробка Высота")
    length_sm = parameters_of_item_in_sm("Коробка Длина")
    WIDTH = parameters_of_item_in_mm(width_sm)
    LENGTH = parameters_of_item_in_mm(length_sm)
    HEIGHT = parameters_of_item_in_mm(height_sm)

    ITEMS = {'article_number': ARTICLE_NUMBER, 'name': NAME, 'price': PRICE,
             'price_without_discount': PRICE_WITHOUT_DISCOUNT,
             'price_ozon_premium': PRICE_OZON_PREMIUM, 'weight': WEIGHT, 'width': WIDTH, 'height': HEIGHT,
             'length': LENGTH, 'article_number_of_photo': ARTICLE_NUMBER_OF_PHOTO, 'brand': BRAND,
             'type_of_item': TYPE_OF_ITEM, 'annotation': ANNOTATION}

    num = list(range(1, (len(ITEMS['article_number'])) + 1))
    for number in range(0, len(num)):
        worksheet.cell(row=number + 4, column=1).value = num[number]

    for i in range(4, len(ITEMS['article_number']) + 4):
        worksheet.cell(row=i, column=3).value = ITEMS['name']
        worksheet.cell(row=i, column=11).value = ITEMS['weight']
        worksheet.cell(row=i, column=19).value = ITEMS['brand']
        worksheet.cell(row=i, column=21).value = ITEMS['type_of_item']

    for r in range(0, len(ITEMS['article_number'])):
        worksheet.cell(row=r + 4, column=2).value = ITEMS['article_number'][r]
        worksheet.cell(row=r + 4, column=4).value = ITEMS['price'][r]
        worksheet.cell(row=r + 4, column=5).value = ITEMS['price_without_discount'][r]
        worksheet.cell(row=r + 4, column=6).value = ITEMS['price_ozon_premium'][r]
        worksheet.cell(row=r + 4, column=15).value = ITEMS['article_number_of_photo'][r]

    for d in range(0, len(ITEMS['width'])):
        worksheet.cell(row=d + 4, column=12).value = ITEMS['width'][d]
        worksheet.cell(row=d + 4, column=13).value = ITEMS['height'][d]
        worksheet.cell(row=d + 4, column=14).value = ITEMS['length'][d]

    for l in range(0, len(ITEMS['annotation'])):
        el = None
        if ITEMS['annotation'][l]:
            el = "\n".join(ITEMS['annotation'][l])
        worksheet.cell(row=l + 4, column=22).value = el

    wb_new.save(NEW_FILE)
